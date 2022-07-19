###################
# Author: Hu Guo
# This program is used to automatically schedule work for members of Champaign Chinese Christian Church on Sunday.
# v0.2003
###################
# 安裝Python的Excel插件
# 在命令行下輸入： pip3 install openpyxl
from calendar import c
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)

# 讀取Excel表格初始可用人員
wb = openpyxl.load_workbook("kai_Excel_2_Macro.xlsm", data_only=True)
# print(type(wb))

sheets = wb.sheetnames
# print(sheets)
# print(wb.active.title)

sh1 = wb['Sheet1']
# print(type(sh1))

####################################################
# 獲取上月已排班兩次人員名單
last_month_assigned_twice = []

curr_year = sh1['B1'].value

curr_month = sh1['E1'].value
print("\n當前安排", curr_year, "年",curr_month, "月份值班人員")

# 生成上月值班表格文件名
if curr_month == 1:
    str_p = str(curr_year - 1) + "年12月擔班情況.xlsx"
else:
    str_p = str(curr_year) + "年" +str(curr_month - 1) + "月擔班情況.xlsx"

try:
    wp = openpyxl.load_workbook(str_p)
    shp_1 = wp['sheet1']
    # print(shp_1.cell(1, 1).value)
    for i in range(1, 25):
        person = shp_1.cell(2, i).value
        ct = shp_1.cell(3, i).value
        if ct == 2:
            # print(name)
            last_month_assigned_twice.append(person)
except:
    print("沒有發現" + str_p)

print("上月擔班了兩次的人員名單: ", last_month_assigned_twice)
print()

# 生成周日具體日期list
sundays = []
for i in range(7, 12):
    date = sh1.cell(3, i).value
    if date != None:
        date2 = str(date)
        dt = datetime.strptime(date2, '%Y-%m-%d %H:%M:%S')
        str_date = dt.strftime('%m/%d')
        if date != None:
            sundays.append(str_date)


# 本月所有週日天數
days = len(sundays)
print("本月有", days, "個禮拜天:")
print("所有週日日期： ", sundays,"\n")

# 建立成員技能名單list
tasks = ["音控主控","音控副控","堂內招待","堂外招待"]
skilled_list = [[],[],[],[]]
for i in range(2, 6):
    for j in range(4, 30):
        person = sh1.cell(j, i).value
        if person != None:
            skilled_list[i - 2].append(sh1.cell(j, 1).value)

print("成員技能名單：")
for i in range(0, len(tasks)):
    print(tasks[i], ": ", skilled_list[i])
print()

# 建立本月可值班人員2D list
available_list = []
for i in range(0, days):
    available_list.append([])
#     available_list[i].append("test")
# available_list[2].append("test")
# print(available_list)

# 建立空白的Dictionary來存儲已安排了項目的人員與值班的天數
attendance_list = dict()

# 從初始表格讀取待可值班人員並加入到本月可值班人員2D list
for i in range(13, 13 + days):
    for j in range(4, 30):
        person = sh1.cell(j, i).value
        if person != None:
            available_list[i - 13].append(sh1.cell(j, i).value)
            attendance_list[person] = 0


# 測試打印初始人員名單
print("本月可安排人員與日期如下：")
for i in range(0, days):
    print(sundays[i], available_list[i])
print()

# 參照上月擔班概要降低上月已經擔班兩次的人員的優先級別
for person in attendance_list:
    if person in last_month_assigned_twice:
        attendance_list[person] = 1
print("本月所有可安排人員如下， 共", len(attendance_list), "人, 如果上月已經擔班兩次，那麼初始優先級會通過 \"+1\" 降低：")
print(attendance_list)
print()

# 建立空白項目分配2D list
arranged_lists = []
for i in range(0, days):
    arranged_lists.append([])
#     arranged_lists[i].append("test")
# arranged_lists[2].append("test")
# print(arranged_lists)
# #############################################################################
for day in range(0, days):
    for task in range(0, len(tasks)):
        attended_0 = [person for person in available_list[day] if attendance_list[person] == 0]
        workers = [person for person in skilled_list[task] if person in attended_0]
        if len(workers) > 0:
            worker = workers[0]
            arranged_lists[day].append(worker)
            attendance_list[worker] = attendance_list.get(worker) + 1
        else:
            attended_1 = [person for person in available_list[day] if attendance_list[person] == 1]
            workers = [person for person in skilled_list[task] if person in attended_1]
            if len(workers) > 0:
                worker = workers[0]
                arranged_lists[day].append(worker)
                attendance_list[worker] = attendance_list.get(worker) + 1
            else:
                worker = "無安排"
                arranged_lists[day].append(worker)
# print(attendance_list, "\n")
# for day in range(0, days):
#     print(arranged_lists[day])

print("本月排班結果： 共", len(sundays), "個週日。")
for day in range(0, days):
    print(sundays[day], arranged_lists[day])
print()

print("根據上月擔班概要和本月值班概要, 增加只擔班一次人員的下月擔班優先級.")
print("如果上月已經擔班兩次，本月在最開始運行程序時已經手動 \"+1\" 降低過擔班優先級。")
print("所以上月擔班兩次的人員在本月實際擔班一次的情況下會顯示值班2次。那麼現在會通過 \"-1\" 增加下月排班的優先級，下月能夠擔班兩次：")
print()

# 根據上月擔班概要和本月值班概要增加只擔班一次人員的下月擔班優先級
for person in attendance_list:
    if person in last_month_assigned_twice:
        attendance_list[person] = attendance_list.get(person) - 1
print("優化完畢")
print()

attended_1 = [person for person in attendance_list if attendance_list.get(person) == 1]
print("本月實際擔班一次的人員有", len(attended_1), "人。 下月可以擔班兩次。")
print(attended_1)
print()

attended_2 = [x for x in attendance_list if attendance_list.get(x) == 2]
print("本月實際擔班兩次的人員有", len(attended_2), "人。 下月盡量只安排一次擔班。")
print(attended_2, "\n")

##########################################################
#  所有成員名單
total_member_list =[]
for row in range(4, 30):
    person = sh1.cell(row, 1).value
    if person != None:
        total_member_list.append(person)
# print(total_member_list)


# 生成優先安排的人名單
total_attendance = dict()

for person in attendance_list:
    total_attendance[person] = []

for row in range(4, 30):
    person = sh1.cell(row, 1).value
    duties = []
    for column in range(7, 7 + days):
        if person not in total_attendance:
            row += 1
        else:
            data = sh1.cell(row, column).value
            if data == None or data == "其他":
                duties.append("")
            else:
                duties.append(data)
    if person in total_attendance:
        total_attendance[person] += duties

# for person in total_attendance:
#     print(person, " ", total_attendance.get(person))

# 本月綜合擔班情況
print("本月綜合擔班情況")
for day in range(0, days):
    for task in range(0, len(tasks)):
        total_attendance[arranged_lists[day][task]][day] = tasks[task]
for person in total_attendance:
    print(person, " ", total_attendance.get(person))

# print(total_attendance)


####################################################
# 保存排班概況與詳細安排到新xlsx文件
wv = Workbook()
wv['Sheet'].title = "sheet1"
shv = wv.active
shv['B1'].value = curr_month
shv['C1'].value = "月擔班次數"
column = 2
for person in attendance_list:
    shv.cell(2, column).value = person
    currentCell = shv.cell(2, column)  # or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='center')
    shv.cell(3, column).value = attendance_list[person]
    shv.cell(4, column).value = "次"
    currentCell = shv.cell(4, column)  # or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='right')
    column += 1

row = 7
for day in sundays:
    shv.cell(row, 1).value = day
    row += 1

shv['B6'].value = "影視主控"
shv['B6'].fill = PatternFill("solid", fgColor="FFC300")
currentCell = shv['B6']
currentCell.alignment = Alignment(horizontal='center')
shv['C6'].value = "影視副控"
shv['C6'].fill = PatternFill("solid", fgColor="FFC300")
currentCell = shv['C6']
currentCell.alignment = Alignment(horizontal='center')
shv['D6'].value = "門口招待"
shv['D6'].fill = PatternFill("solid", fgColor="FFC300")
currentCell = shv['D6']
currentCell.alignment = Alignment(horizontal='center')
shv['E6'].value = "堂內招待"
shv['E6'].fill = PatternFill("solid", fgColor="FFC300")
currentCell = shv['E6']
currentCell.alignment = Alignment(horizontal='center')

for i in range(7, 7 + days):
    for j in range(2, 6):
        shv.cell(i, j).value = arranged_lists[i - 7][j - 2]
        currentCell = shv.cell(i, j)
        currentCell.alignment = Alignment(horizontal='center')
        # if currentCell.value == "藍凱威":
        if currentCell.value == "無安排":
            shv.cell(i, j + 5).value = "空缺建議："
            list_suggest = [person for person in available_list[i - 7] if person not in arranged_lists[i - 7]]
            if len(list_suggest) == 0:
                shv.cell(i, j + 6).value = "無建議"
            else:
                for s in range(0, len(list_suggest)):
                    shv.cell(i, j + 6 + s).value = list_suggest[s]

shv['B13'].value = curr_month
shv['C13'].value = "月綜合擔班次數"
count = 2
for person in total_attendance:
    shv.cell(14, count).value = person
    currentCell = shv.cell(14, count)  # or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='center')
    attendance = 0
    for day in range(0, days):
        if total_attendance[person][day] != "":
            attendance += 1
    shv.cell(15, count).value = attendance
    shv.cell(16, count).value = "次"
    currentCell = shv.cell(16, count)  # or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='right')
    count += 1

row = 17
for day in sundays:
    shv.cell(row, 1).value = day
    row += 1

persons = []
for person in total_attendance:
    persons.append(person)

for column in range(2, 2 + len(total_attendance)):
    # shv.cell(17, column).value = persons[column - 2]
    # currentCell = shv.cell(18, column)
    # currentCell.alignment = Alignment(horizontal='center')
    # # print(total_attendance.get(persons[column - 2]))
    temp = []
    temp = total_attendance.get(persons[column - 2])
    for row in range(17, 17 + days):
        value = temp[row - 17]
        if value != "":
            # print (value)
            shv.cell(row, column).value = value
            currentCell = shv.cell(row, column)
            currentCell.alignment = Alignment(horizontal='center')

str_1 = str(curr_year) + "年" + str(curr_month) + "月擔班情況" + ".xlsx"
wv.save(str_1)
